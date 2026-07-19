# -*- coding: utf-8 -*-
import re
import pandas as pd
from openpyxl import load_workbook

# ============ 精灵球 捕获率/行为 ============
BALLS = {
 'poke-ball':'基础球，捕获率倍率×1。',
 'great-ball':'捕获率倍率×1.5。',
 'ultra-ball':'捕获率倍率×2。',
 'master-ball':'必定捕获成功（捕获率100%）。',
 'safari-ball':'狩猎地带专用，捕获率倍率×1.5。',
 'friend-ball':'捕获的宝可梦亲密度直接变为200。',
 'love-ball':'与同行宝可梦为同一种类且异性时，捕获率倍率×8。',
 'level-ball':'同行宝可梦等级越高捕获率越高（最高约×8）。',
 'lure-ball':'钓鱼时使出捕获率倍率×3（其他方式×1）。',
 'moon-ball':'由月之石进化的宝可梦捕获率倍率×4。',
 'fast-ball':'速度种族值≥100的宝可梦捕获率倍率×4。',
 'heavy-ball':'按体重加算捕获率（最重+30），越重越容易捕获。',
 'sport-ball':'捕虫大会专用，捕获率倍率×1.5。',
 'premier-ball':'捕获率倍率×1，购买精灵球时的赠品。',
 'repeat-ball':'已登入图鉴的宝可梦捕获率倍率×3。',
 'timer-ball':'战斗回合越多捕获率越高（每回合约+0.3，最高×4）。',
 'nest-ball':'目标等级越低捕获率越高（等级≤10时约×3）。',
 'net-ball':'水属性或虫属性宝可梦捕获率倍率×3。',
 'dive-ball':'水面／水下或钓鱼时捕获率倍率×3.5（其他×1）。',
 'luxury-ball':'捕获率倍率×1，但捕获后亲密度提升更快。',
 'heal-ball':'捕获后宝可梦的HP与异常状态完全回复。',
 'quick-ball':'战斗第1回合使出时捕获率倍率×5（其他×1）。',
 'dusk-ball':'洞窟或夜晚使出时捕获率倍率×3.5（其他×1）。',
 'cherish-ball':'活动赠送专用，无法用于捕捉野生宝可梦。',
 'park-ball':'伙伴公园专用，必定捕获成功。',
 'dream-ball':'处于睡眠状态的宝可梦捕获率倍率×4（其他×1）。',
 'beast-ball':'究极异兽捕获率倍率×5（其他×0.1）。',
}

# ============ 回复道具 ============
RECOVERY = {
 'potion':'回复２０点ＨＰ。','super-potion':'回复６０点ＨＰ。','hyper-potion':'回复２００点ＨＰ。',
 'max-potion':'回复全部ＨＰ。','full-restore':'回复全部ＨＰ并治愈所有异常状态。',
 'fresh-water':'回复５０点ＨＰ。','soda-pop':'回复６０点ＨＰ。','lemonade':'回复８０点ＨＰ。',
 'moo-moo-milk':'回复１００点ＨＰ。','energy-powder':'回复６０点ＨＰ（但亲密度下降）。',
 'energy-root':'回复１２０点ＨＰ（但亲密度大幅下降）。','heal-powder':'治愈所有异常状态（但亲密度下降）。',
 'revival-herb':'复活并使ＨＰ回复一半（但亲密度大幅下降）。',
 'antidote':'治愈中毒。','burn-heal':'治愈灼伤。','ice-heal':'治愈冰冻。',
 'awakening':'治愈睡眠。','paralysis-heal':'治愈麻痹。','full-heal':'治愈所有异常状态。',
 'revive':'复活并回复一半ＨＰ。','max-revive':'复活并回复全部ＨＰ。',
 'sacred-ash':'使我方全部失去战斗能力的宝可梦复活并回复全部ＨＰ。',
 'berry-juice':'ＨＰ低于１／２时自动回复２０点ＨＰ（仅一次）。',
 'sweet-heart':'回复２０点ＨＰ。','old-gateau':'治愈所有异常状态。',
 'casteliacone':'回复８０点ＨＰ。','lava-cookie':'回复８０点ＨＰ。','shalour-sable':'回复６０点ＨＰ。',
 'fruit-ball':'回复８０点ＨＰ。','roto-prize-money':'对战后获得的零花钱变为３倍。',
}

# ============ 战斗道具 ============
BATTLE = {
 'x-attack':'战斗中使用１次，攻击提升２级（离场后失效）。',
 'x-defense':'战斗中使用１次，防御提升２级（离场后失效）。',
 'x-speed':'战斗中使用１次，速度提升２级（离场后失效）。',
 'x-accuracy':'战斗中使用１次，命中率提升２级（离场后失效）。',
 'x-sp-atk':'战斗中使用１次，特攻提升２级（离场后失效）。',
 'x-sp-def':'战斗中使用１次，特防提升２级（离场后失效）。',
 'dire-hit':'战斗中使用１次，会心率提升２级（离场后失效）。',
 'guard-spec':'５回合内防止我方能力被降低。',
 'x-attack-2':'战斗中使用１次，攻击提升２级（离场后失效）。',
 'x-defense-2':'战斗中使用１次，防御提升２级（离场后失效）。',
}

# ============ 树果 ============
BERRIES = {
 # 状态解除
 'cheri-berry':'携带时陷入麻痹则自动解除并消耗。','chesto-berry':'携带时陷入睡眠则自动解除并消耗。',
 'pecha-berry':'携带时陷入中毒则自动解除并消耗。','rawst-berry':'携带时陷入灼伤则自动解除并消耗。',
 'aspear-berry':'携带时陷入冰冻则自动解除并消耗。','persim-berry':'携带时陷入生气则自动解除并消耗。',
 'lum-berry':'携带时陷入任何异常状态则自动解除并消耗。','heal-powder':'治愈所有异常状态（但亲密度下降）。',
 # HP/PP
 'leppa-berry':'ＰＰ归零时自动回复１０点ＰＰ并消耗。','oran-berry':'ＨＰ低于１／２时自动回复１０点ＨＰ并消耗。',
 'sitrus-berry':'ＨＰ低于１／２时自动回复最大ＨＰ的１／４并消耗。',
 'hopo-berry':'吃下后可回复１０点ＰＰ。','enigma-berry':'受到效果绝佳伤害时随机获得一种效果。',
 'rowap-berry':'携带时受到特殊招式攻击，给予对手少量伤害并消耗。',
 # 性格果（HP<1/4 提升能力2级）
 'liechi-berry':'ＨＰ≤１／４时攻击提升２级并消耗。','ganlon-berry':'ＨＰ≤１／４时防御提升２级并消耗。',
 'salac-berry':'ＨＰ≤１／４时速度提升２级并消耗。','petaya-berry':'ＨＰ≤１／４时特攻提升２级并消耗。',
 'apicot-berry':'ＨＰ≤１／４时特防提升２级并消耗。','lansat-berry':'ＨＰ≤１／４时会心率提升２级并消耗。',
 'starf-berry':'ＨＰ≤１／４时随机一项能力提升２级并消耗。','micle-berry':'ＨＰ≤１／４时下一只招式命中率提升（不会Miss）并消耗。',
 # 口味果（HP<1/2 回复1/3，讨厌则混乱）
 'figy-berry':'ＨＰ低于１／２时回复最大ＨＰ的１／３；若讨厌其口味则陷入混乱。',
 'wiki-berry':'ＨＰ低于１／２时回复最大ＨＰ的１／３；若讨厌其口味则陷入混乱。',
 'mago-berry':'ＨＰ低于１／２时回复最大ＨＰ的１／３；若讨厌其口味则陷入混乱。',
 'aguav-berry':'ＨＰ低于１／２时回复最大ＨＰ的１／３；若讨厌其口味则陷入混乱。',
 'iapapa-berry':'ＨＰ低于１／２时回复最大ＨＰ的１／３；若讨厌其口味则陷入混乱。',
 # 减弱效果绝佳（伤害减半）
 'occa-berry':'受到效果绝佳的火属性招式时，威力减弱（伤害减半）。','passho-berry':'受到效果绝佳的水属性招式时，威力减弱（伤害减半）。',
 'wacan-berry':'受到效果绝佳的电属性招式时，威力减弱（伤害减半）。','rindo-berry':'受到效果绝佳的草属性招式时，威力减弱（伤害减半）。',
 'yache-berry':'受到效果绝佳的冰属性招式时，威力减弱（伤害减半）。','chople-berry':'受到效果绝佳的格斗属性招式时，威力减弱（伤害减半）。',
 'kebia-berry':'受到效果绝佳的毒属性招式时，威力减弱（伤害减半）。','shuca-berry':'受到效果绝佳的地面属性招式时，威力减弱（伤害减半）。',
 'coba-berry':'受到效果绝佳的飞行属性招式时，威力减弱（伤害减半）。','payapa-berry':'受到效果绝佳的超能属性招式时，威力减弱（伤害减半）。',
 'tanga-berry':'受到效果绝佳的虫属性招式时，威力减弱（伤害减半）。','chilan-berry':'受到效果绝佳的一般属性招式时，威力减弱（伤害减半）。',
 'kasib-berry':'受到效果绝佳的幽灵属性招式时，威力减弱（伤害减半）。','haban-berry':'受到效果绝佳的龙属性招式时，威力减弱（伤害减半）。',
 'colbur-berry':'受到效果绝佳的恶属性招式时，威力减弱（伤害减半）。','babiri-berry':'受到效果绝佳的钢属性招式时，威力减弱（伤害减半）。',
 'roseli-berry':'受到效果绝佳的妖精属性招式时，威力减弱（伤害减半）。',
 # 培育果（提升亲密度但降低某项基础点数）
 'pomeg-berry':'提升亲密度，但ＨＰ基础点数减少。','kelpsy-berry':'提升亲密度，但攻击基础点数减少。',
 'qualot-berry':'提升亲密度，但防御基础点数减少。','hondew-berry':'提升亲密度，但特攻基础点数减少。',
 'grepa-berry':'提升亲密度，但特防基础点数减少。','tamato-berry':'提升亲密度，但速度基础点数减少。',
 # 方块果
 'watmel-berry':'用于制作宝可方块，提升「聪明」造型。','durin-berry':'用于制作宝可方块，提升「美丽」造型。',
 'belue-berry':'用于制作宝可方块，提升「美丽」造型。','spelon-berry':'用于制作宝可方块，提升「可爱」造型。',
 'pamtre-berry':'用于制作宝可方块，提升「可爱」造型。','cornn-berry':'用于制作宝可方块，提升「坚强」造型。',
 'magost-berry':'用于制作宝可方块，提升「可爱」造型。','rabuta-berry':'用于制作宝可方块，提升「坚强」造型。',
 'nomel-berry':'用于制作宝可方块，提升「聪明」造型。','spela-berry':'用于制作宝可方块，提升「坚强」造型。',
 'poffin-case':'（容器，非食用）',
 # 其他
 'berry':'携带时回复自身１０点ＨＰ并消耗。',
}

# ============ 携带物品（重点） ============
HELD = {
 # 通用强化
 'leftovers':'每回合结束时回复最大ＨＰ的１／１６。',
 'black-sludge':'毒属性每回合回复最大ＨＰ的１／１６；非毒属性每回合损失最大ＨＰ的１／８。',
 'shell-bell':'造成伤害的１／８回复自身ＨＰ。',
 'big-root':'吸取类招式回复量提升３０％。',
 # 选择系列
 'choice-band':'攻击提升５０％，但只能使用上场后使出的第一个招式。',
 'choice-specs':'特攻提升５０％，但只能使用上场后使出的第一个招式。',
 'choice-scarf':'速度提升５０％，但只能使用上场后使出的第一个招式。',
 # 强力/弱点
 'life-orb':'招式威力提升３０％，但每回合损失最大ＨＰ的１／１０。',
 'expert-belt':'效果绝佳的招式威力提升２０％。',
 'wise-glasses':'特殊招式威力提升１０％。','muscle-band':'物理招式威力提升１０％。',
 'metronome':'连续使用同一招式时，威力每回合提升１０％（最多２０％）。',
 'type-enhancing' :'',
 # 防御类
 'eviolite':'未进化宝可梦的防御与特防提升５０％。',
 'assault-vest':'特防提升５０％，但只能使用攻击招式。',
 'focus-sash':'满ＨＰ时受到致命伤害仅剩１ＨＰ（每场战斗一次）。',
 'focus-band':'１０％几率承受致命伤害仅剩１ＨＰ。',
 'rocky-helmet':'受到接触类招式攻击时，攻击方受到最大ＨＰ的１／６伤害。',
 'air-balloon':'免疫地面属性招式（被击中后消失）。',
 'safety-goggles':'免疫沙暴、冰雹等天气伤害及粉末／孢子类招式。',
 'heavy-duty-boots':'免疫场地与天气造成的间接伤害。',
 'clear-amulet':'防止自身能力被对手降低。',
 'covert-cloak':'免疫招式的附加效果（仍受伤害）。',
 'loaded-dice':'多段招式的段数倾向更多。',
 'punching-glove':'拳击招式威力提升１０％且不被接触类招式影响。',
 'binding-band':'束缚／绑定造成的伤害提升（每回合约最大ＨＰ的１／８）。',
 'weakness-policy':'受到效果绝佳伤害时，攻击与特攻各提升２级。',
 'red-card':'被击中后强制对手换下一只宝可梦（自身不换）。',
 'eject-button':'被击中后强制自身换下。','eject-pack':'能力被降低时强制自身换下。',
 'mirror-herb':'复制对手获得的能力提升效果。','shed-shell':'换下场时清除自身能力变化与状态。',
 'adrenaline-orb':'被威吓等迫使退场的效果触发时，速度提升１级。',
 # 属性宝石（使用后消耗，该属性招式+30%一次）
 'fire-gem':'火属性招式威力提升３０％（使用一次后消耗）。','water-gem':'水属性招式威力提升３０％（使用一次后消耗）。',
 'grass-gem':'草属性招式威力提升３０％（使用一次后消耗）。','electric-gem':'电属性招式威力提升３０％（使用一次后消耗）。',
 'ice-gem':'冰属性招式威力提升３０％（使用一次后消耗）。','fighting-gem':'格斗属性招式威力提升３０％（使用一次后消耗）。',
 'poison-gem':'毒属性招式威力提升３０％（使用一次后消耗）。','ground-gem':'地面属性招式威力提升３０％（使用一次后消耗）。',
 'flying-gem':'飞行属性招式威力提升３０％（使用一次后消耗）。','psychic-gem':'超能属性招式威力提升３０％（使用一次后消耗）。',
 'bug-gem':'虫属性招式威力提升３０％（使用一次后消耗）。','rock-gem':'岩石属性招式威力提升３０％（使用一次后消耗）。',
 'ghost-gem':'幽灵属性招式威力提升３０％（使用一次后消耗）。','dragon-gem':'龙属性招式威力提升３０％（使用一次后消耗）。',
 'dark-gem':'恶属性招式威力提升３０％（使用一次后消耗）。','steel-gem':'钢属性招式威力提升３０％（使用一次后消耗）。',
 'normal-gem':'一般属性招式威力提升３０％（使用一次后消耗）。','fairy-gem':'妖精属性招式威力提升３０％（使用一次后消耗）。',
 # 属性石板（属性+20%）
 'draco-plate':'龙属性招式威力提升２０％。','dread-plate':'恶属性招式威力提升２０％。',
 'earth-plate':'地面属性招式威力提升２０％。','fairy-plate':'妖精属性招式威力提升２０％。',
 'flame-plate':'火属性招式威力提升２０％。','icicle-plate':'冰属性招式威力提升２０％。',
 'insect-plate':'虫属性招式威力提升２０％。','iron-plate':'钢属性招式威力提升２０％。',
 'meadow-plate':'草属性招式威力提升２０％。','mind-plate':'超能属性招式威力提升２０％。',
 'sky-plate':'飞行属性招式威力提升２０％。','splash-plate':'水属性招式威力提升２０％。',
 'spooky-plate':'幽灵属性招式威力提升２０％。','stone-plate':'岩石属性招式威力提升２０％。',
 'toxic-plate':'毒属性招式威力提升２０％。','zap-plate':'电属性招式威力提升２０％。',
 'pixie-plate':'妖精属性招式威力提升２０％。','spooky-plate-2':'幽灵属性招式威力提升２０％。',
 # 属性强化道具（属性+20%）
 'charcoal':'火属性招式威力提升２０％。','mystic-water':'水属性招式威力提升２０％。',
 'miracle-seed':'草属性招式威力提升２０％。','never-melt-ice':'冰属性招式威力提升２０％。',
 'black-belt':'格斗属性招式威力提升２０％。','poison-barb':'毒属性招式威力提升２０％。',
 'sharp-beak':'飞行属性招式威力提升２０％。','spell-tag':'幽灵属性招式威力提升２０％。',
 'twisted-spoon':'超能属性招式威力提升２０％。','black-glasses':'恶属性招式威力提升２０％。',
 'magnet':'电属性招式威力提升２０％。','dragon-fang':'龙属性招式威力提升２０％。',
 'silk-scarf':'一般属性招式威力提升２０％。','soft-sand':'地面属性招式威力提升２０％。',
 'hard-stone':'岩石属性招式威力提升２０％。','silver-powder':'虫属性招式威力提升２０％。',
 'metal-coat':'钢属性招式威力提升２０％。','silk-scarf-2':'一般属性招式威力提升２０％。',
 # 命中/会心
 'wide-lens':'招式命中率提升１０％。','zoom-lens':'后手时招式命中率提升２０％。',
 'scope-lens':'招式更容易击中要害（会心率提升）。','bright-powder':'对手招式命中率降低１０％。',
 'lax-incense':'对手招式命中率降低１０％。','stick':'大葱鸭携带时招式更容易击中要害。',
 'lucky-punch':'吉利蛋携带时招式更容易击中要害。','razor-fang':'３０％几率使被击中者陷入害怕。',
 'king-s-rock':'３０％几率使被击中者陷入害怕而退缩。',
 # 天气/场地延长
 'damp-rock':'使出的求雨持续时间延长至８回合。','heat-rock':'使出的大晴天持续时间延长至８回合。',
 'icy-rock':'使出的冰雹持续时间延长至８回合。','smooth-rock':'使出的沙暴持续时间延长至８回合。',
 'light-clay':'反射壁／光墙／戏法空间持续时间延长至８回合。',
 # 属性/状态变化
 'flame-orb':'每回合陷入灼伤。','toxic-orb':'每回合陷入剧毒。',
 'light-ball':'皮卡丘携带时特攻提升１００％（×２）。','thick-club':'卡拉卡拉／嘎啦嘎啦携带时攻击提升１００％。',
 'metal-powder':'百变怪携带时防御提升１００％。','quick-powder':'百变怪携带时速度提升１００％。',
 'soul-dew':'拉帝亚斯／拉帝欧斯携带时超能／龙属性招式威力提升２０％。',
 'deep-sea-tooth':'珍珠贝携带时特攻大幅提升（×２）。','deep-sea-scale':'珍珠贝携带时防御大幅提升。',
 'pure-incense':'必定能从野生宝可梦的战斗中逃走。','full-incense':'能力比对手更慢（后手）。',
 'lagging-tail':'能力比对手更慢（后手）。','amulet-coin':'战斗获得金钱变为２倍。',
 'luck-incense':'战斗获得金钱变为２倍。','cleanse-tag':'野生宝可梦出现率降低。',
 'smoke-ball':'必定能从野生战斗中逃走。','chilan-berry-2':'免疫一般属性招式（见树果）。',
 # 努力值道具
 'power-bracer':'携带时获得的攻击努力值翻倍。','power-belt':'携带时获得的防御努力值翻倍。',
 'power-lens':'携带时获得的特攻努力值翻倍。','power-band':'携带时获得的特防努力值翻倍。',
 'power-anklet':'携带时获得的速度努力值翻倍。','power-weight':'携带时获得的ＨＰ努力值翻倍。',
 # 其他
 'wave-incense':'帮助宝可梦放松，用于培育。','rose-incense':'用于培育（产蛋相关）。',
 'sea-incense':'用于培育（产蛋相关）。','odd-incense':'用于培育（产蛋相关）。',
 'rock-incense':'用于培育（产蛋相关）。','sage-charm':'（装饰/无战斗数值效果）。',
 'reaper-cloth':'用于培育（进化相关）。','oval-stone':'用于培育（进化相关）。',
 'twisted-spoon-2':'超能属性招式威力提升２０％。',
 # 重要通用
 'quick-claw':'２５％几率能够先制行动。',
 'everstone':'携带的宝可梦不会进化；孵蛋时５０％几率遗传性格。',
 'lucky-egg':'携带的宝可梦获得经验值提升５０％。',
 'white-herb':'战斗开始时消除自身所有能力下降（每场一次）。',
 'mental-herb':'消除挑衅／力量戏法等带来的束缚效果（每场一次）。',
 'power-herb':'蓄力招式可于使出当回合立即发动。',
 'soothe-bell':'亲密度提升速度加快（约１.５倍）。',
 'destiny-knot':'孵蛋时父母双方各随机遗传５项个体值中的高个体。',
 'protective-pads':'不受接触类招式的附加效果影响。',
 'terrain-extender':'场地类状态持续时间延长至８回合。',
 'float-stone':'携带时体重减半（影响招式威力与先制）。',
 'ring-target':'使对手对自身的招式必定命中（自身更易被击中）。',
 'sticky-barb':'接触类招式攻击后附着对手，每回合损失其最大ＨＰ的１／８；可被夺取。',
 'grip-claw':'束缚／绑紧等招式的持续时间延长至最大值。',
 'razor-claw':'招式更容易击中要害（会心率提升）；用于特定进化。',
 'iron-ball':'速度减半且免疫地面属性招式；用于特定进化。',
 'fist-plate':'格斗属性招式威力提升２０％。',
 'adamant-orb':'帝牙卢卡携带时钢／龙属性招式威力提升２０％。',
 'lustrous-orb':'帕路奇亚携带时水／龙属性招式威力提升２０％。',
 'griseous-orb':'骑拉帝纳携带时幽灵／龙属性招式威力提升２０％。',
 'red-orb':'固拉多携带时进入原始回归形态。','blue-orb':'盖欧卡携带时进入原始回归形态。',
 'macho-brace':'努力值获得翻倍，但速度减半。',
 'blunder-policy':'招式落空（Miss）时速度提升２级。',
 'cell-battery':'受到电属性招式攻击时特攻提升１级并消耗。',
 'absorb-bulb':'受到水属性招式攻击时特攻提升１级并消耗。',
 'electric-seed':'电气场地出现时防御提升（并消耗）。','psychic-seed':'精神场地出现时特防提升（并消耗）。',
 'misty-seed':'薄雾场地出现时特防提升（并消耗）。','grassy-seed':'青草场地出现时防御提升（并消耗）。',
 'berserk-gene':'攻击提升２级，但之后必定陷入混乱（一次性）。',
 'dragon-scale':'用于培育（暴飞龙进化道具）。',
 'protector':'用于培育（特定宝可梦进化道具）。',
 'electirizer':'用于培育（电击兽进化道具）。','magmarizer':'用于培育（鸭嘴炎兽进化道具）。',
 'dubious-disc':'用于培育（多边兽进化道具）。','prism-scale':'用于培育（丑丑鱼进化道具）。',
 'reaper-cloth':'用于培育（魔墙人偶进化道具）。','upgrade':'用于培育（多边兽进化道具）。',
 'whipped-dream':'用于培育（粉香香进化道具）。','sachet':'用于培育（绵绵泡芙进化道具）。',
 'luminous-moss':'用于培育（opy? 提升亲密度道具）。','snowball':'用于培育（穿山鼠进化道具）。',
 'pink-bow':'装饰用，旧作中提升一般属性招式威力。','polkadot-bow':'装饰用，旧作中提升一般属性招式威力。',
 'red-scarf':'华丽大赛装饰，提升酷炫表现。','blue-scarf':'华丽大赛装饰，提升美丽表现。',
 'pink-scarf':'华丽大赛装饰，提升可爱表现。','green-scarf':'华丽大赛装饰，提升聪明表现。',
 'yellow-scarf':'华丽大赛装饰，提升坚强表现。',
 'douse-drive':'代欧奇希斯携带时改变属性／形态。','shock-drive':'代欧奇希斯携带时改变属性／形态。',
 'burn-drive':'代欧奇希斯携带时改变属性／形态。','chill-drive':'代欧奇希斯携带时改变属性／形态。',
}
# 旧版树果（第一/二代残留）
BERRIES.update({
 'przcureberry':'吃下后治愈麻痹。','psncureberry':'吃下后治愈中毒。',
 'burnt-berry':'吃下后治愈灼伤。','ice-berry':'吃下后治愈冰冻。',
 'bitter-berry':'吃下后治愈混乱。','miracleberry':'吃下后治愈所有异常状态。',
 'mysteryberry':'吃下后回复少量ＰＰ。','gold-berry':'旧版树果，回复少量ＨＰ。',
})

# 属性强化通用（部分英文名为 xxx-plate / xxx-gem 已在上面映射）
# 补充：第八世代对战道具 / 战斗树果 / 传说球
BERRIES.update({
 'charti-berry':'受到效果绝佳的岩石属性招式时，威力减弱（伤害减半）。',
 'custap-berry':'ＨＰ≤１／４时必定先制行动并消耗。',
 'jaboca-berry':'受到物理招式攻击时给予对手伤害并消耗。',
 'kee-berry':'受到物理招式攻击时防御提升１级并消耗。',
 'maranga-berry':'受到特殊招式攻击时特防提升１级并消耗。',
})
HELD.update({
 'rusted-sword':'苍响携带时进入剑之王形态（属性变为钢／妖精）。',
 'rusted-shield':'藏玛然特携带时进入盾之王形态（属性变为格斗／岩石）。',
 'fairy-feather':'妖精属性招式威力提升２０％。',
 'booster-energy':'对应的悖论宝可梦携带时，依天气提升速度或防御／特防。',
 'ability-shield':'防止自身特性被更改或消除。',
 'throat-spray':'使出变化招式后特攻提升１级（每场一次）。',
 'room-service':'处于下雨／沙暴等天气时速度降低（对手相对更快）。',
 'utility-umbrella':'免疫天气造成的自身能力强化／弱化。',
 'leader-s-crest':'古简蜗携带时提升特定招式威力。',
 'cornerstone-mask':'厄诡椪携带时进入础石面具形态。',
 'wellspring-mask':'厄诡椪携带时进入水井面具形态。',
 'hearthflame-mask':'厄诡椪携带时进入火灶面具形态。',
 'adamant-crystal':'帝牙卢卡携带时钢／龙属性招式威力提升２０％。',
 'lustrous-globe':'帕路奇亚携带时水／龙属性招式威力提升２０％。',
 'griseous-core':'骑拉帝纳携带时幽灵／龙属性招式威力提升２０％。',
 'up-grade':'用于培育（多边兽进化道具）。',
 'mint-berry':'旧版树果，回复少量ＨＰ。','snowball-2':'用于培育（穿山鼠进化道具）。',
})
BALLS.update({
 'leaden-ball':'对体重重的宝可梦捕获率更高。',
 'gigaton-ball':'对体重极大的宝可梦捕获率极高（投掷较慢）。',
 'feather-ball':'对飞行属性宝可梦捕获率更高，且更容易命中。',
 'wing-ball':'对飞行属性宝可梦捕获率更高。',
 'jet-ball':'对速度极快的宝可梦捕获率更高。',
 'strange-ball':'捕获率随宝可梦种类而变化（特殊球）。',
 'razz-berry':'宝可GO中用于捕捉，略微提升捕获率。',
 'pinap-berry':'宝可GO中使捕捉后获得的糖果翻倍。',
 'nanab-berry':'宝可GO中用于捕捉，使宝可梦不易行动。',
 'silver-razz-berry':'宝可GO中提升捕获率（更强）。','golden-razz-berry':'宝可GO中大幅提升捕获率。',
 'bluk-berry':'宝可GO中用于制作莓莓果酱（捕捉辅助）。','wepear-berry':'宝可GO中用于制作莓莓果酱（捕捉辅助）。',
 'silver-nanab-berry':'宝可GO中使宝可梦更不易行动。','golden-nanab-berry':'宝可GO中使宝可梦更不易行动。',
 'silver-pinap-berry':'宝可GO中糖果翻倍且提升捕获率。','golden-pinap-berry':'宝可GO中糖果大幅翻倍。',
})

def get_effect(slug, zh, cat, desc, cur):
    slug = str(slug).strip()
    if cat == '精灵球':
        return BALLS.get(slug, '用于捕捉野生宝可梦的球。')
    if cat == '回复道具':
        if slug in RECOVERY: return RECOVERY[slug]
        if cur and re.search(r'[０-９0-9％%／/]', str(cur)): return cur
        return desc
    if cat == '战斗道具':
        return BATTLE.get(slug, desc)
    if cat == '树果':
        return BERRIES.get(slug, desc)
    if cat == '携带物品':
        if slug in HELD and HELD[slug]: return HELD[slug]
        # 存储碟 (Silvally 属性变化)
        if slug.endswith('-memory'):
            t = slug[:-7]
            zh = {'fighting':'格斗','flying':'飞行','poison':'毒','ground':'地面','rock':'岩石',
                  'bug':'虫','ghost':'幽灵','steel':'钢','fire':'火','water':'水','grass':'草',
                  'electric':'电','psychic':'超能','ice':'冰','dragon':'龙','dark':'恶','normal':'一般',
                  'fairy':'妖精'}.get(t, t)
            return f'黏龙携带时属性变为{zh}属性。'
        if slug.endswith('-drive'):
            return '代欧奇希斯携带时改变属性／形态。'
        return desc
    if cat == '超级石':
        name = zh.replace('进化石','')
        return f'携带后，{name}在战斗中可以超级进化。'
    if cat in ('训练家使用的Ｚ纯晶','宝可梦使用的Ｚ纯晶'):
        attr = zh.replace('Ｚ','').strip()
        return f'携带后，可将{attr}属性的招式转化为对应的Ｚ招式（威力大幅提升并附带额外效果）。'
    if cat == '洛托姆之力':
        return cur if (cur and cur != 'nan') else desc
    return cur if cur else desc

if __name__ == '__main__':
    wb = load_workbook('宝可梦小图鉴_数据.xlsx')
    ws = wb['道具']
    hdr = [c.value for c in ws[1]]
    ci = {n:i+1 for i,n in enumerate(hdr)}
    keep = ['携带物品','树果','回复道具','战斗道具','超级石','训练家使用的Ｚ纯晶','宝可梦使用的Ｚ纯晶','洛托姆之力','精灵球']
    # 删除非对战分类的行（从下往上）
    del_rows=[]
    for r in range(2, ws.max_row+1):
        c = ws.cell(r, ci['分类']).value
        if c not in keep:
            del_rows.append(r)
    for r in reversed(del_rows):
        ws.delete_rows(r, 1)
    # 重新统计行
    filled=0; total=ws.max_row-1
    unknown=[]
    for r in range(2, ws.max_row+1):
        slug = ws.cell(r, ci['英文名']).value
        zh = ws.cell(r, ci['中文名']).value
        cat = ws.cell(r, ci['分类']).value
        desc = ws.cell(r, ci['描述']).value
        cur = ws.cell(r, ci['效果']).value
        eff = get_effect(slug, zh, cat, desc, cur)
        ws.cell(r, ci['效果']).value = eff
        filled+=1
        # 标记未能具体化的
        if (cat in ('携带物品','树果','精灵球') and (slug not in HELD and slug not in BERRIES and slug not in BALLS)):
            unknown.append((slug, zh, cat))
    wb.save('宝可梦小图鉴_数据.xlsx')
    print(f'道具保留行数: {total}')
    print(f'已重写效果: {filled}')
    print(f'未精确匹配(回退到描述)的携带/树果/球: {len(unknown)}')
    for u in unknown[:60]: print('   ', u)

    # ===== 招式 补充描述列 =====
    mvd = pd.read_excel('宝可梦小图鉴_数据.xlsx', sheet_name='招式描述')
    desc_map = {}
    for _,r in mvd.iterrows():
        en=r['英文名']; d=r['描述']
        if pd.notna(en) and pd.notna(d):
            desc_map[str(en).strip()] = str(d).strip()
    ws2 = wb['招式']
    h2 = [c.value for c in ws2[1]]
    if '描述' not in h2:
        nc = len(h2)+1
        ws2.cell(1, nc, '描述')
        added=0
        for r in range(2, ws2.max_row+1):
            en = ws2.cell(r, 1).value
            if en is not None and str(en).strip() in desc_map:
                ws2.cell(r, nc, desc_map[str(en).strip()]); added+=1
        print(f'\n招式 描述列已添加: {added} 行')
    else:
        print('\n招式 已含描述列，跳过')
    wb.save('宝可梦小图鉴_数据.xlsx')
    print('全部已保存。')
